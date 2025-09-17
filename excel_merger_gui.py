import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import pandas as pd
import os
import glob
import shutil
from pathlib import Path
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import threading

class ExcelMergerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("고급 엑셀 파일 통합기 v2.0")
        self.root.geometry("800x600")
        
        # 파일 리스트와 작업 폴더
        self.excel_files = []
        self.class_info_file = None
        self.work_folder = None
        
        self.setup_ui()
        
    def setup_ui(self):
        """UI 구성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 제목
        title_label = ttk.Label(main_frame, text="🚀 고급 엑셀 파일 통합기", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # 작업 폴더 선택
        folder_frame = ttk.LabelFrame(main_frame, text="작업 폴더 설정", padding="10")
        folder_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Button(folder_frame, text="작업 폴더 선택", 
                  command=self.select_work_folder).grid(row=0, column=0, padx=(0, 10))
        
        self.folder_label = ttk.Label(folder_frame, text="폴더를 선택해주세요", 
                                     foreground="gray")
        self.folder_label.grid(row=0, column=1, sticky=tk.W)
        
        # 파일 목록 프레임
        file_frame = ttk.LabelFrame(main_frame, text="엑셀 파일 목록", padding="10")
        file_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 파일 추가 버튼들
        button_frame = ttk.Frame(file_frame)
        button_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Button(button_frame, text="파일 추가", 
                  command=self.add_files).grid(row=0, column=0, padx=(0, 5))
        ttk.Button(button_frame, text="수업정보 파일 선택", 
                  command=self.select_class_info).grid(row=0, column=1, padx=5)
        ttk.Button(button_frame, text="목록 초기화", 
                  command=self.clear_files).grid(row=0, column=2, padx=5)
        
        # 드래그 앤 드롭 영역
        self.drop_frame = tk.Frame(file_frame, bg="lightgray", height=100)
        self.drop_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        self.drop_frame.grid_propagate(False)
        
        drop_label = tk.Label(self.drop_frame, text="여기에 엑셀 파일을 드래그하세요", 
                             bg="lightgray", font=('Arial', 10))
        drop_label.pack(expand=True)
        
        # 드래그 앤 드롭 이벤트 바인딩
        self.drop_frame.drop_target_register("DND_Files")
        self.drop_frame.dnd_bind("<<Drop>>", self.on_drop)
        
        # 파일 목록 표시
        columns = ("파일명", "크기", "상태")
        self.file_tree = ttk.Treeview(file_frame, columns=columns, show="headings", height=8)
        
        for col in columns:
            self.file_tree.heading(col, text=col)
            self.file_tree.column(col, width=200)
        
        self.file_tree.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(file_frame, orient=tk.VERTICAL, command=self.file_tree.yview)
        scrollbar.grid(row=2, column=2, sticky=(tk.N, tk.S))
        self.file_tree.configure(yscrollcommand=scrollbar.set)
        
        # 수업정보 파일 표시
        self.class_info_label = ttk.Label(file_frame, text="수업정보 파일: 선택되지 않음", 
                                         foreground="orange")
        self.class_info_label.grid(row=3, column=0, columnspan=3, sticky=tk.W, pady=(10, 0))
        
        # 실행 버튼
        execute_frame = ttk.Frame(main_frame)
        execute_frame.grid(row=3, column=0, columnspan=3, pady=20)
        
        self.execute_button = ttk.Button(execute_frame, text="🔄 파일 통합 실행", 
                                        command=self.execute_merge, style="Accent.TButton")
        self.execute_button.pack()
        
        # 진행상황 표시
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 로그 출력 영역
        log_frame = ttk.LabelFrame(main_frame, text="처리 로그", padding="5")
        log_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.log_text = ScrolledText(log_frame, height=10, state='disabled')
        self.log_text.pack(fill='both', expand=True)
        
        # 그리드 웨이트 설정
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        file_frame.columnconfigure(0, weight=1)
        file_frame.rowconfigure(2, weight=1)
        
    def log(self, message):
        """로그 메시지 출력"""
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.config(state='disabled')
        self.log_text.see(tk.END)
        self.root.update()
        
    def select_work_folder(self):
        """작업 폴더 선택"""
        folder = filedialog.askdirectory()
        if folder:
            self.work_folder = folder
            self.folder_label.config(text=folder, foreground="black")
            self.log(f"작업 폴더 설정: {folder}")
            
    def add_files(self):
        """파일 추가"""
        files = filedialog.askopenfilenames(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        for file_path in files:
            self.add_file_to_list(file_path)
            
    def select_class_info(self):
        """수업정보 파일 선택"""
        file_path = filedialog.askopenfilename(
            title="수업정보 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            self.class_info_file = file_path
            filename = os.path.basename(file_path)
            self.class_info_label.config(text=f"수업정보 파일: {filename}", foreground="green")
            self.log(f"수업정보 파일 선택: {filename}")
            
    def add_file_to_list(self, file_path):
        """파일을 목록에 추가 (중복 체크)"""
        filename = os.path.basename(file_path)
        
        # 중복 체크
        for item in self.file_tree.get_children():
            existing_filename = self.file_tree.item(item)['values'][0]
            if existing_filename == filename:
                self.log(f"중복 파일 건너뜀: {filename}")
                return
        
        # 파일 크기 계산
        try:
            size = os.path.getsize(file_path)
            size_str = f"{size:,} bytes"
:
            size_str = "알 수 없음"
        
        # 트리뷰에 추가
        item_id = self.file_tree.insert("", tk.END, values=(filename, size_str, "대기"))
        self.excel_files.append(file_path)
        self.log(f"파일 추가: {filename}")
        
    def on_drop(self, event):
        """드래그 앤 드롭 이벤트 처리"""
        files = self.root.tk.splitlist(event.data)
        for file_path in files:
            if file_path.lower().endswith(('.xlsx', '.xls')):
                self.add_file_to_list(file_path)
            else:
                self.log(f"지원하지 않는 파일 형식: {os.path.basename(file_path)}")
                
    def clear_files(self):
        """파일 목록 초기화"""
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        self.excel_files = []
        self.log("파일 목록이 초기화되었습니다.")
        
    def update_file_status(self, filename, status):
        """파일 상태 업데이트"""
        for item in self.file_tree.get_children():
            if self.file_tree.item(item)['values'][0] == filename:
                values = list(self.file_tree.item(item)['values'])
                values[2] = status
                self.file_tree.item(item, values=values)
                break
                
    def execute_merge(self):
        """파일 통합 실행"""
        if not self.work_folder:
            messagebox.showerror("오류", "작업 폴더를 선택해주세요.")
            return
            
        if not self.excel_files:
            messagebox.showerror("오류", "통합할 엑셀 파일을 추가해주세요.")
            return
            
        # 별도 스레드에서 실행
        thread = threading.Thread(target=self._execute_merge_thread)
        thread.daemon = True
        thread.start()
        
    def _execute_merge_thread(self):
        """파일 통합 실행 (스레드)"""
        try:
            self.execute_button.config(state='disabled')
            self.progress.start(10)
            
            # 작업 폴더 권한 체크
            try:
                test_file = os.path.join(self.work_folder, "test_permission.tmp")
                with open(test_file, 'w') as f:
                    f.write("test")
                os.remove(test_file)
                self.log(f"✅ 작업 폴더 권한 확인: {self.work_folder}")
            except Exception as e:
                self.log(f"❌ 작업 폴더 권한 없음: {e}")
                messagebox.showerror("권한 오류", f"작업 폴더에 쓰기 권한이 없습니다.\n다른 폴더를 선택해주세요.\n{self.work_folder}")
                return
            
            # 작업 폴더에 파일 복사
            self.log("작업 폴더로 파일 복사 중...")
            copied_files = []
            
            for file_path in self.excel_files:
                filename = os.path.basename(file_path)
                dest_path = os.path.join(self.work_folder, filename)
                
                try:
                    # 파일이 열려있는지 체크
                    if os.path.exists(dest_path):
                        try:
                            # 파일을 쓰기 모드로 열어서 사용 중인지 확인
                            with open(dest_path, 'r+b'):
                                pass
                        except PermissionError:
                            self.log(f"⚠️ 파일이 사용 중입니다. 닫고 다시 시도하세요: {filename}")
                            self.update_file_status(filename, "사용중")
                            continue
                    
                    # 중복 파일 체크 및 복사
                    if os.path.exists(dest_path):
                        if os.path.getsize(file_path) == os.path.getsize(dest_path):
                            self.log(f"동일한 파일 존재, 건너뜀: {filename}")
                            self.update_file_status(filename, "건너뜀")
                            copied_files.append(dest_path)
                            continue
                    
                    # 파일 복사 전 원본 파일 체크
                    if not os.path.exists(file_path):
                        self.log(f"❌ 원본 파일이 존재하지 않음: {filename}")
                        self.update_file_status(filename, "원본없음")
                        continue
                    
                    shutil.copy2(file_path, dest_path)
                    copied_files.append(dest_path)
                    self.update_file_status(filename, "복사완료")
                    self.log(f"파일 복사: {filename}")
                    
                except PermissionError as e:
                    self.log(f"권한 오류 - {filename}: {e}")
                    self.update_file_status(filename, "권한오류")
                except Exception as e:
                    self.log(f"파일 복사 실패 {filename}: {e}")
                    self.update_file_status(filename, "복사실패")
            
            if not copied_files:
                self.log("❌ 복사된 파일이 없습니다!")
                messagebox.showerror("오류", "복사된 파일이 없어 통합을 진행할 수 없습니다.")
                return
            
            # 수업정보 파일 복사
            class_info_path = None
            if self.class_info_file:
                class_info_filename = os.path.basename(self.class_info_file)
                class_info_path = os.path.join(self.work_folder, class_info_filename)
                try:
                    if not os.path.exists(self.class_info_file):
                        self.log(f"⚠️ 수업정보 파일이 존재하지 않음: {class_info_filename}")
                    else:
                        shutil.copy2(self.class_info_file, class_info_path)
                        self.log(f"수업정보 파일 복사: {class_info_filename}")
                except Exception as e:
                    self.log(f"수업정보 파일 복사 실패: {e}")
                    class_info_path = None
            
            # 파일 통합 실행
            self.log("파일 통합 시작...")
            success = self.merge_excel_files_advanced(self.work_folder, class_info_path)
            
            if success:
                self.log("✅ 파일 통합 완료!")
                
                # 결과 파일 확인
                output_file = os.path.join(self.work_folder, "통합파일_고급.xlsx")
                if os.path.exists(output_file):
                    file_size = os.path.getsize(output_file)
                    self.log(f"📄 결과 파일: 통합파일_고급.xlsx ({file_size:,} bytes)")
                    messagebox.showinfo("완료", f"파일 통합이 성공적으로 완료되었습니다!\n\n저장 위치: {output_file}")
                else:
                    self.log("⚠️ 결과 파일이 생성되지 않았습니다.")
                    messagebox.showwarning("경고", "통합은 완료되었지만 결과 파일을 찾을 수 없습니다.")
                
                # 모든 파일 상태를 완료로 변경
                for item in self.file_tree.get_children():
                    values = list(self.file_tree.item(item)['values'])
                    if values[2] in ["복사완료", "건너뜀"]:
                        values[2] = "완료"
                        self.file_tree.item(item, values=values)
            else:
                self.log("❌ 파일 통합 실패!")
                messagebox.showerror("오류", "파일 통합에 실패했습니다. 로그를 확인해주세요.")
                
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.log(f"심각한 오류: {e}")
            self.log(f"오류 상세: {error_detail}")
            messagebox.showerror("오류", f"예상치 못한 오류가 발생했습니다:\n{e}\n\n상세 로그를 확인해주세요.")
            
        finally:
            self.progress.stop()
            self.execute_button.config(state='normal')

    # 기존 엑셀 처리 함수들을 여기에 포함
    def extract_education_name_and_date(self, df):
        """첫 번째 행에서 교육명과 날짜를 추출"""
        if df.empty or len(df) < 1:
            return None, None, None
        
        first_row = df.iloc[0]
        education_name = None
        date_str = None
        time_str = None
        
        for cell_value in first_row:
            if pd.isna(cell_value):
                continue
                
            cell_str = str(cell_value).strip()
            
            # 교육명 패턴 찾기
            if not education_name and re.search(r'[가-힣]', cell_str) and len(cell_str) > 2:
                education_name = cell_str
            
            # 날짜 패턴 찾기
            date_patterns = [
                r'(\d{1,2})[-/.](\d{1,2})',
                r'(\d{1,2})월\s*(\d{1,2})일',
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, cell_str)
                if match and not date_str:
                    month, day = match.groups()
                    date_str = f"{month.zfill(2)}-{day.zfill(2)}"
                    break
            
            # 시간 패턴 찾기
            time_match = re.search(r'(\d{1,2}):(\d{2})', cell_str)
            if time_match and not time_str:
                hour, minute = time_match.groups()
                time_str = f"{hour.zfill(2)}-{minute}"
        
        return education_name, date_str, time_str

    def clean_sheet_name(self, name):
        """엑셀 시트명으로 사용할 수 없는 문자들 정리"""
        if not name:
            return "Sheet"
        
        invalid_chars = r'[\[\]:*?/\\]'
        cleaned = re.sub(invalid_chars, '_', str(name))
        
        if len(cleaned) > 31:
            cleaned = cleaned[:28] + "..."
        
        if not cleaned.strip():
            cleaned = "Sheet"
        
        return cleaned

    def process_sheet_data(self, df):
        """시트 데이터 처리"""
        if df.empty or len(df) < 3:
            return None
        
        # 첫 번째와 두 번째 행 삭제
        df = df.iloc[2:].reset_index(drop=True)
        
        if df.empty:
            return None
        
        # 컬럼명 정리
        if len(df) > 0:
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)
        
        # 빈 행 제거
        df = df.dropna(how='all')
        
        if df.empty:
            return None
        
        # 취소 행과 일반 행 분리
        cancelled_rows = []
        normal_rows = []
        
        for idx, row in df.iterrows():
            is_cancelled = False
            for col in df.columns:
                if pd.notna(row[col]) and '취소' in str(row[col]):
                    is_cancelled = True
                    break
            
            if is_cancelled:
                cancelled_rows.append(row)
            else:
                normal_rows.append(row)
        
        df_normal = pd.DataFrame(normal_rows).reset_index(drop=True) if normal_rows else pd.DataFrame()
        df_cancelled = pd.DataFrame(cancelled_rows).reset_index(drop=True) if cancelled_rows else pd.DataFrame()
        
        return df_normal, df_cancelled

    def update_waitlist_status(self, df):
        """대기 학생 상태 업데이트"""
        if df.empty:
            return df
        
        status_col = None
        for col in df.columns:
            if '상태' in str(col) or 'status' in str(col).lower():
                status_col = col
                break
        
        if status_col is None:
            return df
        
        for idx, row in df.iterrows():
            status_value = str(row[status_col]) if pd.notna(row[status_col]) else ""
            
            wait_match = re.search(r'대기(\d+)', status_value)
            if wait_match:
                wait_num = wait_match.group(1)
                df.at[idx, status_col] = f"Applied (대기{wait_num})"
        
        return df

    def sort_dataframe(self, df):
        """데이터프레임 정렬"""
        if df.empty:
            return df
        
        sort_columns = []
        sort_orders = []
        
        column_mapping = {
            '상태': ('상태', False),
            '지역': ('지역', True),
            '학교분류': ('학교분류', False),
            '학교명': ('학교명', True),
            '학년': ('학년', True),
            '반': ('반', True),
            '번호': ('번호', True)
        }
        
        for col_name, (ascending, is_asc) in column_mapping.items():
            for df_col in df.columns:
                if col_name in str(df_col):
                    sort_columns.append(df_col)
                    sort_orders.append(is_asc)
                    break
        
        if sort_columns:
            df = df.sort_values(by=sort_columns, ascending=sort_orders, na_position='last')
        
        return df

    def add_grade_class_number_column(self, df):
        """학년-반-번호 조합 열 추가"""
        if df.empty:
            return df
        
        grade_col = None
        class_col = None
        number_col = None
        
        for col in df.columns:
            col_str = str(col)
            if '학년' in col_str:
                grade_col = col
            elif '반' in col_str and '학년' not in col_str:
                class_col = col
            elif '번호' in col_str:
                number_col = col
        
        if grade_col and class_col and number_col:
            def combine_grade_class_number(row):
                grade = str(row[grade_col]) if pd.notna(row[grade_col]) else ""
                class_num = str(row[class_col]) if pd.notna(row[class_col]) else ""
                number = str(row[number_col]) if pd.notna(row[number_col]) else ""
                
                if grade and class_num and number:
                    return f"{grade}-{class_num}-{number}"
                return ""
            
            df['학년-반-번호'] = df.apply(combine_grade_class_number, axis=1)
        
        return df

    def add_class_info_columns(self, df, class_info_df, education_name):
        """수업 정보 열 추가"""
        if df.empty:
            return df
        
        # 기본 열들 추가
        new_columns = ['수업일', '시작', '종료', '주강사', '보조강사', '장소', '모니터']
        for col in new_columns:
            df[col] = ""
        
        # 수업 정보가 있으면 매핑
        if not class_info_df.empty and education_name:
            # 교육명으로 해당 수업 정보 찾기
            matching_rows = class_info_df[class_info_df.apply(
                lambda row: any(education_name in str(cell) for cell in row), axis=1
            )]
            
            if not matching_rows.empty:
                class_info = matching_rows.iloc[0]
                
                # 컬럼 매핑
                column_mapping = {
                    '수업일': ['수업일', '날짜', 'date', '일자'],
                    '시작': ['시작', '시작시간', 'start', '시작시간'],
                    '종료': ['종료', '종료시간', 'end', '끝'],
                    '주강사': ['주강사', '강사', 'instructor', '선생님'],
                    '보조강사': ['보조강사', '보조', 'assistant'],
                    '장소': ['장소', '위치', 'location', '교실'],
                    '모니터': ['모니터', 'monitor', '관리자']
                }
                
                for target_col, possible_cols in column_mapping.items():
                    for possible_col in possible_cols:
                        for class_col in class_info_df.columns:
                            if possible_col in str(class_col).lower():
                                value = str(class_info[class_col]) if pd.notna(class_info[class_col]) else ""
                                df[target_col] = value
                                break
                        if df[target_col].iloc[0]:  # 값이 설정되었으면 다음으로
                            break
        
        return df

    def apply_strikethrough_to_cancelled_rows(self, worksheet, cancelled_df, start_row):
        """취소된 행에 취소선 적용"""
        if cancelled_df.empty:
            return
        
        for idx, row in cancelled_df.iterrows():
            row_num = start_row + idx + 1
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_idx)
                cell.font = Font(strike=True)
                cell.value = value

    def merge_excel_files_advanced(self, folder_path, class_info_file=None, output_file='통합파일_고급.xlsx'):
        """고급 엑셀 파일 통합"""
        try:
            # 엑셀 파일들 찾기
            excel_files = []
            for ext in ['*.xlsx', '*.xls']:
                excel_files.extend(glob.glob(os.path.join(folder_path, ext)))
            
            output_path = os.path.join(folder_path, output_file)
            excel_files = [f for f in excel_files if os.path.abspath(f) != os.path.abspath(output_path)]
            if class_info_file:
                excel_files = [f for f in excel_files if os.path.abspath(f) != os.path.abspath(class_info_file)]
            
            if not excel_files:
                self.log("❌ 폴더에 처리할 엑셀 파일이 없습니다.")
                return False
            
            self.log(f"📂 발견된 엑셀 파일: {len(excel_files)}개")
            for file_path in excel_files:
                self.log(f"   - {os.path.basename(file_path)}")
            
            # 수업 정보 파일 읽기
            class_info_df = pd.DataFrame()
            if class_info_file and os.path.exists(class_info_file):
                try:
                    # 여러 시트가 있을 수 있으므로 첫 번째 시트 읽기
                    class_info_df = pd.read_excel(class_info_file, sheet_name=0)
                    self.log(f"📋 수업 정보 파일 로드: {os.path.basename(class_info_file)} ({len(class_info_df)}행)")
                except Exception as e:
                    self.log(f"⚠️ 수업 정보 파일 읽기 실패: {e}")
            
            # 교육명별로 그룹화하여 처리
            education_groups = {}
            processed_count = 0
            
            for file_path in excel_files:
                try:
                    filename = os.path.basename(file_path)
                    self.log(f"📄 처리 중: {filename}")
                    
                    # 파일이 열려있는지 체크
                    try:
                        with open(file_path, 'r+b'):
                            pass
                    except PermissionError:
                        self.log(f"⚠️ 파일이 사용 중입니다. 건너뜀: {filename}")
                        continue
                    
                    # 엑셀 파일 읽기
                    try:
                        excel_data = pd.ExcelFile(file_path)
                    except Exception as e:
                        self.log(f"❌ 엑셀 파일 읽기 실패: {filename} - {e}")
                        continue
                    
                    sheet_processed = 0
                    for sheet_name in excel_data.sheet_names:
                        try:
                            # 시트 읽기 (헤더 없이)
                            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                            
                            if df.empty or len(df) < 3:
                                self.log(f"  ⚠️ 데이터가 부족한 시트: {sheet_name}")
                                continue
                            
                            # 교육명과 날짜 추출
                            education_name, date_str, time_str = self.extract_education_name_and_date(df)
                            
                            if not education_name:
                                self.log(f"  ⚠️ 교육명을 찾을 수 없음: {sheet_name}")
                                continue
                            
                            # 시트 데이터 처리
                            result = self.process_sheet_data(df)
                            if result is None:
                                self.log(f"  ⚠️ 처리할 데이터가 없음: {sheet_name}")
                                continue
                            
                            df_normal, df_cancelled = result
                            
                            if df_normal.empty:
                                self.log(f"  ⚠️ 유효한 데이터가 없음: {sheet_name}")
                                continue
                            
                            # 데이터 처리
                            df_normal = self.update_waitlist_status(df_normal)
                            df_normal = self.sort_dataframe(df_normal)
                            df_normal = self.add_grade_class_number_column(df_normal)
                            df_normal = self.add_class_info_columns(df_normal, class_info_df, education_name)
                            
                            # 시트명 생성
                            if date_str and time_str:
                                sheet_name_new = f"{education_name}_{date_str}_{time_str}"
                            elif date_str:
                                sheet_name_new = f"{education_name}_{date_str}"
                            else:
                                sheet_name_new = education_name
                            
                            sheet_name_new = self.clean_sheet_name(sheet_name_new)
                            
                            # 교육명별로 그룹화
                            if education_name not in education_groups:
                                education_groups[education_name] = []
                            
                            education_groups[education_name].append({
                                'sheet_name': sheet_name_new,
                                'data': df_normal,
                                'cancelled_data': df_cancelled,
                                'date_str': date_str,
                                'time_str': time_str,
                                'original_file': filename
                            })
                            
                            self.log(f"  ✅ {sheet_name} → {sheet_name_new} ({len(df_normal):,}행)")
                            sheet_processed += 1
                            
                        except Exception as e:
                            self.log(f"  ❌ 시트 처리 오류: {sheet_name} - {e}")
                            continue
                    
                    if sheet_processed > 0:
                        processed_count += 1
                        self.log(f"  📊 {filename}: {sheet_processed}개 시트 처리됨")
                    
                except Exception as e:
                    self.log(f"❌ 파일 처리 오류: {filename} - {e}")
                    continue
            
            if processed_count == 0:
                self.log("❌ 처리된 파일이 없습니다!")
                return False
            
            self.log(f"✅ 총 {processed_count}개 파일에서 데이터 추출 완료")
            
            # 각 교육명별로 최신 버전만 선택하고 통합
            final_sheets = {}
            
            for education_name, sheets in education_groups.items():
                if not sheets:
                    continue
                
                # 날짜별로 그룹화
                date_groups = {}
                for sheet_info in sheets:
                    date_key = sheet_info['date_str'] or 'no_date'
                    if date_key not in date_groups:
                        date_groups[date_key] = []
                    date_groups[date_key].append(sheet_info)
                
                # 각 날짜별로 최신 버전 선택
                for date_key, date_sheets in date_groups.items():
                    if len(date_sheets) == 1:
                        final_sheets[date_sheets[0]['sheet_name']] = date_sheets[0]
                    else:
                        for i, sheet_info in enumerate(date_sheets, 1):
                            if i > 1:
                                sheet_info['sheet_name'] = f"{sheet_info['sheet_name']}*{i-1}"
                            final_sheets[sheet_info['sheet_name']] = sheet_info
            
            if not final_sheets:
                self.log("❌ 최종 생성할 시트가 없습니다!")
                return False
            
            self.log(f"📝 최종 생성될 시트: {len(final_sheets)}개")
            
            # 기존 출력 파일이 열려있는지 체크
            if os.path.exists(output_path):
                try:
                    with open(output_path, 'r+b'):
                        pass
                except PermissionError:
                    self.log(f"❌ 출력 파일이 사용 중입니다: {output_file}")
                    return False
            
            # 엑셀 파일 생성
            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for sheet_name, sheet_info in final_sheets.items():
                        try:
                            # 취소된 데이터와 일반 데이터 합치기
                            if not sheet_info['cancelled_data'].empty:
                                combined_data = pd.concat([
                                    sheet_info['data'],
                                    sheet_info['cancelled_data']
                                ], ignore_index=True)
                            else:
                                combined_data = sheet_info['data']
                            
                            combined_data.to_excel(writer, sheet_name=sheet_name, index=False)
                            self.log(f"  📄 시트 생성: {sheet_name}")
                            
                        except Exception as e:
                            self.log(f"❌ 시트 생성 실패: {sheet_name} - {e}")
                            continue
                
                self.log("✅ 엑셀 파일 생성 완료")
                
                # 취소선 적용
                try:
                    wb = openpyxl.load_workbook(output_path)
                    
                    for sheet_name, sheet_info in final_sheets.items():
                        if not sheet_info['cancelled_data'].empty and sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            normal_rows = len(sheet_info['data'])
                            self.apply_strikethrough_to_cancelled_rows(ws, sheet_info['cancelled_data'], normal_rows + 1)
                    
                    wb.save(output_path)
                    self.log("✅ 취소선 적용 완료")
                    
                except Exception as e:
                    self.log(f"⚠️ 취소선 적용 실패: {e}")
                
            except Exception as e:
                self.log(f"❌ 엑셀 파일 생성 실패: {e}")
                return False
            
            self.log(f"🎉 통합 완료!")
            self.log(f"📊 처리 결과:")
            self.log(f"   - 처리된 파일: {processed_count}개")
            self.log(f"   - 통합된 시트: {len(final_sheets)}개")
            self.log(f"   - 저장 위치: {output_path}")
            
            return True
            
        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.log(f"❌ 프로그램 오류: {e}")
            self.log(f"상세 오류: {error_detail}")
            return False Exception as e:
            self.log(f"❌ 프로그램 오류: {e}")
            return False


def main():
    """메인 실행 함수"""
    try:
        # tkinterdnd2 임포트 (드래그 앤 드롭 기능)
        try:
            from tkinterdnd2 import TkinterDnD, DND_FILES
            root = TkinterDnD.Tk()
        except ImportError:
            print("tkinterdnd2 라이브러리가 설치되지 않았습니다.")
            print("다음 명령어로 설치하세요: pip install tkinterdnd2")
            root = tk.Tk()
        
        app = ExcelMergerGUI(root)
        
        # 드래그 앤 드롭 기능 추가 (tkinterdnd2가 있는 경우)
        if 'TkinterDnD' in str(type(root)):
            def drop_enter(event):
                event.widget.config(bg="lightblue")
                
            def drop_leave(event):
                event.widget.config(bg="lightgray")
                
            def drop(event):
                event.widget.config(bg="lightgray")
                files = root.tk.splitlist(event.data)
                for file_path in files:
                    if file_path.lower().endswith(('.xlsx', '.xls')):
                        app.add_file_to_list(file_path)
                    else:
                        app.log(f"지원하지 않는 파일 형식: {os.path.basename(file_path)}")
            
            app.drop_frame.drop_target_register(DND_FILES)
            app.drop_frame.dnd_bind('<<DropEnter>>', drop_enter)
            app.drop_frame.dnd_bind('<<DropLeave>>', drop_leave)
            app.drop_frame.dnd_bind('<<Drop>>', drop)
        
        # 시작 메시지
        app.log("=" * 50)
        app.log("🚀 고급 엑셀 파일 통합기 v2.0")
        app.log("📧 제작자: charmleader@gmail.com")
        app.log("=" * 50)
        app.log("사용 방법:")
        app.log("1. 작업 폴더를 선택하세요")
        app.log("2. 엑셀 파일들을 추가하거나 드래그하세요")
        app.log("3. (선택사항) 수업정보 파일을 선택하세요")
        app.log("4. '파일 통합 실행' 버튼을 클릭하세요")
        app.log("=" * 50)
        
        root.mainloop()
        
    except Exception as e:
        print(f"프로그램 시작 오류: {e}")
        input("Enter 키를 눌러 종료하세요...")


if __name__ == "__main__":
    main()