import streamlit as st
import pandas as pd
import os
import glob
from pathlib import Path
import re
from datetime import datetime
import warnings
import zipfile
import tempfile
import io
import base64
import shutil
import subprocess
import sys

# pandas 경고 메시지 숨기기
warnings.filterwarnings('ignore')

# 페이지 설정
st.set_page_config(
    page_title="엑셀 통합기 - charmleader.com",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS 스타일
st.markdown("""
<style>
    .header-style {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
    }
    .success-box {
        background: linear-gradient(90deg, #56ab2f 0%, #a8e6cf 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        margin: 1rem 0;
    }
    .info-box {
        background: linear-gradient(90deg, #3498db 0%, #85c1e9 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        margin: 1rem 0;
    }
    .warning-box {
        background: linear-gradient(90deg, #f39c12 0%, #f7dc6f 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        margin: 1rem 0;
    }
    .file-uploader {
        border: 2px dashed #667eea;
        border-radius: 10px;
        padding: 2rem;
        text-align: center;
        background: #f8f9ff;
    }
    .stButton > button {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 25px;
        padding: 0.5rem 2rem;
        font-weight: bold;
        transition: all 0.3s ease;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
    }
</style>
""", unsafe_allow_html=True)

def get_download_link(df, filename, file_label):
    """다운로드 링크 생성"""
    csv = df.to_csv(index=False, encoding='utf-8-sig')
    b64 = base64.b64encode(csv.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="{filename}">📥 {file_label} 다운로드</a>'
    return href

def get_excel_download_link(df, filename, file_label):
    """엑셀 다운로드 링크 생성"""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='통합데이터')
    
    b64 = base64.b64encode(output.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">📊 {file_label} 다운로드</a>'
    return href

def safe_read_excel(file, sheet_name=None):
    """안전한 Excel 파일 읽기 (스타일 오류 방지)"""
    try:
        # 방법 1: openpyxl 엔진 사용
        if sheet_name:
            return pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
        else:
            return pd.read_excel(file, engine='openpyxl')
    except Exception as e1:
        try:
            # 방법 2: xlrd 엔진 사용 (구형 Excel 파일용)
            if sheet_name:
                return pd.read_excel(file, sheet_name=sheet_name, engine='xlrd')
            else:
                return pd.read_excel(file, engine='xlrd')
        except Exception as e2:
            try:
                # 방법 3: 기본 엔진 사용
                if sheet_name:
                    return pd.read_excel(file, sheet_name=sheet_name, engine=None)
                else:
                    return pd.read_excel(file, engine=None)
            except Exception as e3:
                # 방법 4: 스타일 정보 완전 무시하고 읽기
                import openpyxl
                from io import BytesIO
                
                # 파일을 메모리로 읽기
                file.seek(0)
                file_bytes = file.read()
                file.seek(0)  # 파일 포인터 원위치
                
                # openpyxl로 워크북 열기 (데이터만)
                wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                
                if sheet_name and sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.active
                
                # 워크시트를 데이터프레임으로 변환
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])  # 첫 행을 헤더로 사용
                    return df
                else:
                    raise Exception("빈 워크시트")
                    
            except Exception as e4:
                raise Exception(f"모든 읽기 방법 실패: openpyxl({e1}), xlrd({e2}), 기본({e3}), 수동({e4})")

def clean_dataframe(df):
    """데이터프레임 정리"""
    # 빈 행 제거
    df = df.dropna(how='all')
    
    # 빈 열 제거
    df = df.dropna(axis=1, how='all')
    
    # 문자열 컬럼의 앞뒤 공백 제거
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip()
    
    return df

def check_and_request_permissions():
    """시스템 권한 확인 및 요청"""
    try:
        # 사용자 문서 폴더에 쓰기 권한 확인
        documents_path = Path.home() / "Documents"
        test_file = documents_path / "test_permission.tmp"
        
        try:
            # 쓰기 권한 테스트
            test_file.write_text("test")
            test_file.unlink()  # 테스트 파일 삭제
            return True, "권한 확인 완료"
        except PermissionError:
            return False, "문서 폴더에 쓰기 권한이 없습니다. 관리자 권한으로 실행해주세요."
        except Exception as e:
            return False, f"권한 확인 중 오류: {str(e)}"
    
    except Exception as e:
        return False, f"권한 확인 실패: {str(e)}"

def save_uploaded_files(uploaded_files):
    """업로드된 파일들을 사용자 문서 위치에 저장"""
    try:
        # 권한 확인
        has_permission, permission_msg = check_and_request_permissions()
        
        if not has_permission:
            st.warning(f"⚠️ {permission_msg}")
            
            # 권한 요청 안내
            st.markdown("""
            <div class="warning-box">
            <h4>🔐 시스템 권한이 필요합니다</h4>
            <p>Excel 파일을 정리하려면 다음 중 하나를 선택하세요:</p>
            <ol>
                <li><strong>관리자 권한으로 실행:</strong> 프로그램을 마우스 우클릭 → "관리자 권한으로 실행"</li>
                <li><strong>다른 폴더 사용:</strong> 권한이 있는 다른 폴더를 선택하세요</li>
            </ol>
            </div>
            """, unsafe_allow_html=True)
            
            # 대안 폴더 선택
            alt_folder = st.text_input("대안 폴더 경로 (예: C:\\temp):", value=str(Path.home() / "Desktop"))
            
            if st.button("📁 대안 폴더 사용"):
                try:
                    alt_path = Path(alt_folder)
                    alt_path.mkdir(exist_ok=True)
                    excel_merger_folder = alt_path / "엑셀통합기_업로드파일"
                    excel_merger_folder.mkdir(exist_ok=True)
                except Exception as e:
                    st.error(f"대안 폴더 생성 실패: {e}")
                    return None, [], []
            else:
                return None, [], []
        else:
            # 기본 문서 폴더 사용
            documents_path = Path.home() / "Documents"
            excel_merger_folder = documents_path / "엑셀통합기_업로드파일"
            excel_merger_folder.mkdir(exist_ok=True)
        
        # 타임스탬프로 하위 폴더 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        session_folder = excel_merger_folder / f"세션_{timestamp}"
        session_folder.mkdir(exist_ok=True)
        
        saved_files = []
        duplicate_files = []
        
        for uploaded_file in uploaded_files:
            # 파일 저장
            file_path = session_folder / uploaded_file.name
            
            # 중복 파일 확인
            if file_path.exists():
                # 파일명에 타임스탬프 추가
                name_parts = file_path.stem, file_path.suffix
                file_path = session_folder / f"{name_parts[0]}_{timestamp}{name_parts[1]}"
                duplicate_files.append(uploaded_file.name)
            
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            saved_files.append(str(file_path))
        
        return str(session_folder), saved_files, duplicate_files
    
    except Exception as e:
        st.error(f"❌ 파일 저장 중 오류 발생: {str(e)}")
        return None, [], []

def run_launcher_py():
    """launcher.py 실행"""
    try:
        # launcher.py 파일 경로
        launcher_path = "launcher.py"
        
        if os.path.exists(launcher_path):
            # Python으로 launcher.py 실행
            result = subprocess.run([sys.executable, launcher_path], 
                                  capture_output=True, text=True, timeout=30)
            
            if result.returncode == 0:
                return True, "launcher.py가 성공적으로 실행되었습니다."
            else:
                return False, f"launcher.py 실행 실패: {result.stderr}"
        else:
            return False, "launcher.py 파일을 찾을 수 없습니다."
    
    except subprocess.TimeoutExpired:
        return False, "launcher.py 실행 시간이 초과되었습니다."
    except Exception as e:
        return False, f"launcher.py 실행 중 오류: {str(e)}"

def merge_excel_files(uploaded_files, merge_option, sheet_name=None):
    """엑셀 파일들을 통합"""
    all_data = []
    file_info = []
    
    try:
        for uploaded_file in uploaded_files:
            file_name = uploaded_file.name
            
            # 파일 확장자에 따라 읽기
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                try:
                    # 안전한 Excel 읽기 함수 사용
                    df = safe_read_excel(uploaded_file, sheet_name)
                except Exception as e:
                    st.warning(f"⚠️ {file_name} 읽기 실패: {str(e)}")
                    continue
            elif file_name.endswith('.csv'):
                df = pd.read_csv(uploaded_file, encoding='utf-8-sig')
            else:
                st.warning(f"⚠️ 지원하지 않는 파일 형식: {file_name}")
                continue
            
            # 데이터 정리
            df = clean_dataframe(df)
            
            if merge_option == "파일명 추가":
                df['파일명'] = file_name
            elif merge_option == "폴더명 추가":
                df['폴더명'] = os.path.dirname(file_name) or "루트"
            
            all_data.append(df)
            file_info.append({
                '파일명': file_name,
                '행수': len(df),
                '열수': len(df.columns)
            })
    
    except Exception as e:
        st.error(f"❌ 파일 처리 중 오류 발생: {str(e)}")
        return None, None
    
    if not all_data:
        st.error("❌ 처리할 수 있는 파일이 없습니다.")
        return None, None
    
    # 데이터 통합
    try:
        if merge_option == "단순 통합":
            merged_df = pd.concat(all_data, ignore_index=True)
        else:
            merged_df = pd.concat(all_data, ignore_index=True)
        
        # 최종 정리
        merged_df = clean_dataframe(merged_df)
        
        return merged_df, file_info
    
    except Exception as e:
        st.error(f"❌ 데이터 통합 중 오류 발생: {str(e)}")
        return None, None

def main():
    # 헤더
    st.markdown("""
    <div class="header-style">
        <h1>🚀 스마트 엑셀 파일 통합기</h1>
        <h3>🌐 제작자: charmleader.com</h3>
        <p>⏰ 실행 시간: {}</p>
    </div>
    """.format(datetime.now().strftime("%Y-%m-%d %H:%M:%S")), unsafe_allow_html=True)

    # 사이드바
    with st.sidebar:
        st.markdown("### 📋 사용법")
        st.markdown("""
        1. **파일 업로드**: 통합할 엑셀/CSV 파일들을 선택
        2. **통합 옵션**: 원하는 통합 방식 선택
        3. **시트 선택**: 특정 시트만 통합하려면 선택
        4. **통합 실행**: '파일 통합하기' 버튼 클릭
        5. **다운로드**: 통합된 파일 다운로드
        """)
        
        st.markdown("### 📊 지원 형식")
        st.markdown("""
        - **.xlsx** (Excel 2007+)
        - **.xls** (Excel 97-2003)
        - **.csv** (쉼표 구분)
        """)
        
        st.markdown("### ⚙️ 통합 옵션")
        st.markdown("""
        - **단순 통합**: 모든 데이터를 하나로 합치기
        - **파일명 추가**: 각 행에 원본 파일명 추가
        - **폴더명 추가**: 각 행에 원본 폴더명 추가
        """)

    # 메인 컨텐츠
    st.markdown("### 📁 엑셀 파일 업로드")
    
    # 드래그 앤 드롭 스타일 추가
    st.markdown("""
    <style>
    .upload-container {
        border: 3px dashed #667eea;
        border-radius: 15px;
        padding: 40px;
        text-align: center;
        background: #f8f9ff;
        transition: all 0.3s ease;
        margin: 20px 0;
    }
    .upload-container:hover {
        border-color: #764ba2;
        background: #f0f0ff;
    }
    .upload-container.dragover {
        border-color: #28a745;
        background: #f0fff0;
    }
    .upload-icon {
        font-size: 3em;
        color: #667eea;
        margin-bottom: 15px;
    }
    .upload-text {
        font-size: 1.2em;
        color: #495057;
        margin-bottom: 10px;
    }
    .upload-hint {
        color: #6c757d;
        font-size: 0.9em;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 드래그 앤 드롭 영역
    st.markdown("""
    <div class="upload-container" id="upload-container">
        <div class="upload-icon">📁</div>
        <div class="upload-text">엑셀/CSV 파일을 드래그하거나 클릭하여 선택하세요</div>
        <div class="upload-hint">여러 파일을 동시에 선택할 수 있습니다 (.xlsx, .xls, .csv)</div>
    </div>
    """, unsafe_allow_html=True)
    
    # 파일 업로더
    uploaded_files = st.file_uploader(
        "통합할 엑셀/CSV 파일들을 선택하세요",
        type=['xlsx', 'xls', 'csv'],
        accept_multiple_files=True,
        help="여러 파일을 동시에 선택할 수 있습니다",
        label_visibility="collapsed"
    )
    
    if uploaded_files:
        st.markdown(f"<div class='info-box'>📁 총 {len(uploaded_files)}개의 파일이 선택되었습니다.</div>", unsafe_allow_html=True)
        
        # 업로드된 파일 목록 표시
        st.markdown("### 📋 업로드된 파일 목록")
        for i, file in enumerate(uploaded_files, 1):
            file_size = len(file.getbuffer())
            size_mb = file_size / (1024 * 1024)
            st.markdown(f"**{i}.** {file.name} ({size_mb:.2f} MB)")
        
        # 통합 옵션
        col1, col2 = st.columns(2)
        
        with col1:
            merge_option = st.selectbox(
                "통합 방식 선택",
                ["단순 통합", "파일명 추가", "폴더명 추가"],
                help="데이터를 어떻게 통합할지 선택하세요"
            )
        
        with col2:
            # 시트 선택 옵션 (첫 번째 파일의 시트 목록 가져오기)
            sheet_options = ["모든 시트 (첫 번째 시트만)"]
            if uploaded_files:
                try:
                    first_file = uploaded_files[0]
                    if first_file.name.endswith(('.xlsx', '.xls')):
                        # 안전한 방법으로 시트 목록 가져오기
                        import openpyxl
                        from io import BytesIO
                        
                        first_file.seek(0)
                        file_bytes = first_file.read()
                        first_file.seek(0)
                        
                        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                        sheet_options.extend(wb.sheetnames)
                except Exception as e:
                    st.warning(f"시트 목록을 가져올 수 없습니다: {e}")
                    pass
            
            sheet_name = st.selectbox(
                "시트 선택",
                sheet_options,
                help="특정 시트만 통합하려면 선택하세요"
            )
            
            if sheet_name == "모든 시트 (첫 번째 시트만)":
                sheet_name = None
        
        # 파일 저장 및 launcher.py 실행 버튼
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("💾 파일 저장하기", type="secondary"):
                with st.spinner("파일을 저장하는 중..."):
                    session_folder, saved_files, duplicate_files = save_uploaded_files(uploaded_files)
                    
                    if session_folder:
                        st.markdown(f"<div class='success-box'>✅ 파일이 저장되었습니다!</div>", unsafe_allow_html=True)
                        st.markdown(f"**저장 위치:** `{session_folder}`")
                        st.markdown(f"**저장된 파일 수:** {len(saved_files)}개")
                        
                        if duplicate_files:
                            st.markdown(f"**중복 파일 처리:** {len(duplicate_files)}개 파일에 타임스탬프 추가")
                        
                        # 세션 상태에 저장
                        st.session_state['saved_files'] = saved_files
                        st.session_state['session_folder'] = session_folder
        
        with col2:
            if st.button("🚀 파일 통합 실행", type="primary"):
                # 먼저 파일 저장
                with st.spinner("파일을 저장하고 통합하는 중..."):
                    session_folder, saved_files, duplicate_files = save_uploaded_files(uploaded_files)
                    
                    if session_folder:
                        st.markdown(f"<div class='info-box'>📁 파일이 저장되었습니다: {session_folder}</div>", unsafe_allow_html=True)
                        
                        if duplicate_files:
                            st.markdown(f"**중복 파일 처리:** {len(duplicate_files)}개 파일에 타임스탬프 추가")
                        
                        # 파일 통합 실행 (launcher.py 실행 제거)
                        merged_df, file_info = merge_excel_files(
                            uploaded_files, 
                            merge_option, 
                            sheet_name
                        )
                        
                        if merged_df is not None:
                            st.markdown("<div class='success-box'>✅ 파일 통합이 완료되었습니다!</div>", unsafe_allow_html=True)
                            
                            # 파일 정보 표시
                            st.markdown("### 📊 파일 정보")
                            info_df = pd.DataFrame(file_info)
                            st.dataframe(info_df, use_container_width=True)
                            
                            # 통합 결과 미리보기
                            st.markdown("### 👀 통합 결과 미리보기")
                            st.dataframe(merged_df.head(10), use_container_width=True)
                            
                            # 통합 통계
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("총 행수", f"{len(merged_df):,}")
                            with col2:
                                st.metric("총 열수", f"{len(merged_df.columns):,}")
                            with col3:
                                st.metric("처리된 파일수", len(file_info))
                            
                            # 결과 파일 저장 및 폴더 열기
                            try:
                                # 결과 파일 저장
                                result_filename = f"통합데이터_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                result_path = Path(session_folder) / result_filename
                                
                                with pd.ExcelWriter(result_path, engine='openpyxl') as writer:
                                    merged_df.to_excel(writer, index=False, sheet_name='통합데이터')
                                
                                st.markdown(f"<div class='success-box'>💾 결과 파일이 저장되었습니다: {result_path}</div>", unsafe_allow_html=True)
                                
                                # 폴더 열기 버튼
                                if st.button("📁 결과 폴더 열기", type="secondary"):
                                    try:
                                        if sys.platform.startswith('win'):
                                            os.startfile(session_folder)
                                        elif sys.platform.startswith('darwin'):
                                            subprocess.run(['open', session_folder])
                                        else:
                                            subprocess.run(['xdg-open', session_folder])
                                        st.success("폴더가 열렸습니다!")
                                    except Exception as e:
                                        st.error(f"폴더 열기 실패: {e}")
                                
                            except Exception as e:
                                st.error(f"파일 저장 중 오류: {e}")
                            
                            # 다운로드 섹션
                            st.markdown("### 📥 다운로드")
                            
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                # CSV 다운로드
                                csv_filename = f"통합데이터_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                                csv_link = get_download_link(merged_df, csv_filename, "CSV 파일")
                                st.markdown(csv_link, unsafe_allow_html=True)
                            
                            with col2:
                                # Excel 다운로드
                                excel_filename = f"통합데이터_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                excel_link = get_excel_download_link(merged_df, excel_filename, "Excel 파일")
                                st.markdown(excel_link, unsafe_allow_html=True)
                            
                            # 세션 상태에 저장
                            st.session_state['merged_data'] = merged_df
                            st.session_state['file_info'] = file_info
                            st.session_state['result_folder'] = session_folder
    else:
        st.markdown("""
        <div class="file-uploader">
            <h3>📁 파일을 업로드하세요</h3>
            <p>위의 파일 선택 영역을 클릭하여 엑셀 또는 CSV 파일들을 선택하세요</p>
            <p>여러 파일을 동시에 선택할 수 있습니다</p>
        </div>
        """, unsafe_allow_html=True)
    
    # 결과 표시 섹션 (세션 상태에 데이터가 있을 때)
    if 'merged_data' in st.session_state and 'result_folder' in st.session_state:
        st.markdown("---")
        st.markdown("### 📊 최근 처리 결과")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("총 행수", f"{len(st.session_state['merged_data']):,}")
        with col2:
            st.metric("총 열수", f"{len(st.session_state['merged_data'].columns):,}")
        with col3:
            st.metric("처리된 파일수", len(st.session_state['file_info']))
        
        # 결과 폴더 열기 버튼
        if st.button("📁 결과 폴더 열기", type="secondary"):
            try:
                result_folder = st.session_state['result_folder']
                if sys.platform.startswith('win'):
                    os.startfile(result_folder)
                elif sys.platform.startswith('darwin'):
                    subprocess.run(['open', result_folder])
                else:
                    subprocess.run(['xdg-open', result_folder])
                st.success("폴더가 열렸습니다!")
            except Exception as e:
                st.error(f"폴더 열기 실패: {e}")
        
        # 결과 미리보기
        with st.expander("👀 결과 미리보기", expanded=False):
            st.dataframe(st.session_state['merged_data'].head(20), use_container_width=True)
    
    # 푸터
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #666; padding: 2rem;">
        <p>🌐 <strong>charmleader.com</strong> | 💌 문의 및 피드백 언제든 환영합니다!</p>
        <p>⚡ 빠르고 안전한 엑셀 파일 통합 솔루션</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
