import pandas as pd
import os
import glob
from pathlib import Path
import re
from datetime import datetime
import shutil
import stat
import sys

def check_file_permissions(file_path):
    """파일 권한 확인"""
    try:
        # 파일 존재 확인
        if not os.path.exists(file_path):
            return False, "파일이 존재하지 않습니다"
        
        # 읽기 권한 확인
        if not os.access(file_path, os.R_OK):
            return False, "파일 읽기 권한이 없습니다"
        
        # 파일이 잠겨있는지 확인 (다른 프로그램에서 사용 중)
        try:
            with open(file_path, 'rb') as f:
                f.read(1)  # 1바이트만 읽어서 파일이 열리는지 확인
        except PermissionError:
            return False, "파일이 다른 프로그램에서 사용 중입니다"
        except Exception as e:
            return False, f"파일 접근 오류: {e}"
        
        return True, "권한 확인 완료"
    
    except Exception as e:
        return False, f"권한 확인 중 오류: {e}"

def fix_file_permissions(file_path):
    """파일 권한 수정 시도"""
    try:
        # Windows에서 파일 속성 변경
        if sys.platform.startswith('win'):
            # 읽기 전용 속성 제거
            os.chmod(file_path, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IWGRP | stat.S_IROTH)
            
            # 파일을 임시로 복사해서 권한 문제 해결 시도
            temp_path = file_path + ".temp"
            try:
                shutil.copy2(file_path, temp_path)
                shutil.move(temp_path, file_path)
                return True, "파일 권한이 수정되었습니다"
            except Exception:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                return False, "파일 권한 수정에 실패했습니다"
        else:
            # Unix/Linux에서 권한 설정
            os.chmod(file_path, 0o644)
            return True, "파일 권한이 수정되었습니다"
    
    except Exception as e:
        return False, f"권한 수정 중 오류: {e}"

def safe_read_excel(file_path, sheet_name=None):
    """안전한 Excel 파일 읽기 (권한 및 오류 처리)"""
    try:
        # 1단계: 권한 확인
        has_permission, permission_msg = check_file_permissions(file_path)
        if not has_permission:
            print(f"  ⚠️ 권한 문제: {permission_msg}")
            
            # 권한 수정 시도
            fixed, fix_msg = fix_file_permissions(file_path)
            if fixed:
                print(f"  🔧 권한 수정: {fix_msg}")
            else:
                print(f"  ❌ 권한 수정 실패: {fix_msg}")
                return None
        
        # 2단계: Excel 파일 읽기 시도 (스타일 오류 방지)
        try:
            # openpyxl 엔진으로 읽기 (스타일 무시)
            if sheet_name:
                return pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            else:
                return pd.read_excel(file_path, engine='openpyxl')
        except Exception as e1:
            print(f"  ⚠️ openpyxl 오류: {e1}")
            try:
                # xlrd 엔진으로 재시도
                if sheet_name:
                    return pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
                else:
                    return pd.read_excel(file_path, engine='xlrd')
            except Exception as e2:
                print(f"  ⚠️ xlrd 오류: {e2}")
                try:
                    # 기본 엔진으로 재시도
                    if sheet_name:
                        return pd.read_excel(file_path, sheet_name=sheet_name, engine=None)
                    else:
                        return pd.read_excel(file_path, engine=None)
                except Exception as e3:
                    print(f"  ⚠️ 기본 엔진 오류: {e3}")
                    # 수동으로 openpyxl로 읽기 (스타일 완전 무시)
                    try:
                        import openpyxl
                        from io import BytesIO
                        
                        # 파일을 메모리로 읽기
                        with open(file_path, 'rb') as f:
                            file_bytes = f.read()
                        
                        # openpyxl로 워크북 열기 (스타일 무시)
                        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                        
                        if sheet_name and sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                        else:
                            ws = wb.active
                        
                        # 워크시트를 데이터프레임으로 변환
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            data.append(row)
                        
                        if data:
                            # 첫 번째 행을 헤더로 사용
                            if len(data) > 1:
                                df = pd.DataFrame(data[1:], columns=data[0])
                            else:
                                df = pd.DataFrame(data)
                            return df
                        else:
                            raise Exception("빈 워크시트")
                    
                    except Exception as e4:
                        print(f"  ❌ 수동 읽기 실패: {e4}")
                        # 마지막 시도: 파일을 복사해서 읽기
                        try:
                            import tempfile
                            import shutil
                            
                            # 임시 파일로 복사
                            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                                shutil.copy2(file_path, temp_file.name)
                                temp_path = temp_file.name
                            
                            # 복사된 파일로 읽기
                            if sheet_name:
                                df = pd.read_excel(temp_path, sheet_name=sheet_name, engine='openpyxl')
                            else:
                                df = pd.read_excel(temp_path, engine='openpyxl')
                            
                            # 임시 파일 삭제
                            os.unlink(temp_path)
                            return df
                            
                        except Exception as e5:
                            print(f"  ❌ 모든 읽기 방법 실패: {e5}")
                            return None
    
    except Exception as e:
        print(f"  ❌ 파일 읽기 중 심각한 오류: {e}")
        return None

def get_base_filename_and_version(filename):
    """
    파일명에서 기본명과 버전을 분리
    예: "1-1.xlsx" -> ("1", "1"), "report_v2.xlsx" -> ("report", "2")
    """
    # 확장자 제거
    name_without_ext = Path(filename).stem
    
    # 버전 패턴들과 해당 버전 번호 추출
    version_patterns = [
        (r'^(.+)-(\d+)$', 2),      # file-1, file-2
        (r'^(.+)_v(\d+)$', 2),     # file_v1, file_v2
        (r'^(.+)_ver(\d+)$', 2),   # file_ver1, file_ver2
        (r'^(.+)\((\d+)\)$', 2),   # file(1), file(2)
        (r'^(.+)_(\d+)$', 2),      # file_1, file_2
        (r'^(.+)-v(\d+)$', 2),     # file-v1, file-v2
    ]
    
    for pattern, version_group in version_patterns:
        match = re.match(pattern, name_without_ext)
        if match:
            base_name = match.group(1)
            version = match.group(version_group)
            return base_name, version
    
    # 버전 정보가 없으면 원본 이름과 버전 "0"
    return name_without_ext, "0"

def find_latest_versions(excel_files):
    """
    같은 기본 파일명을 가진 파일들 중 최신 버전만 선택
    """
    file_groups = {}
    
    # 기본 파일명별로 그룹화
    for file_path in excel_files:
        filename = os.path.basename(file_path)
        base_name, version = get_base_filename_and_version(filename)
        
        if base_name not in file_groups:
            file_groups[base_name] = []
        
        file_groups[base_name].append({
            'path': file_path,
            'filename': filename,
            'version': int(version) if version.isdigit() else 0,
            'version_str': version
        })
    
    # 각 그룹에서 최신 버전만 선택
    latest_files = []
    version_info = []
    
    for base_name, files in file_groups.items():
        # 버전 번호가 가장 높은 파일 선택
        latest_file = max(files, key=lambda x: x['version'])
        latest_files.append(latest_file)
        
        if len(files) > 1:
            # 여러 버전이 있는 경우
            old_versions = [f['filename'] for f in files if f['version'] < latest_file['version']]
            version_info.append({
                'base_name': base_name,
                'latest': latest_file['filename'],
                'latest_version': latest_file['version_str'],
                'old_versions': old_versions
            })
        else:
            # 단일 버전
            version_info.append({
                'base_name': base_name,
                'latest': latest_file['filename'],
                'latest_version': latest_file['version_str'],
                'old_versions': []
            })
    
    return latest_files, version_info

def clean_sheet_name(name):
    """엑셀 시트명으로 사용할 수 없는 문자들 정리"""
    if not name:
        return "Sheet"
    
    # 엑셀에서 허용하지 않는 문자들 제거: [ ] : * ? / \
    invalid_chars = r'[\[\]:*?/\\]'
    cleaned = re.sub(invalid_chars, '_', str(name))
    
    # 시트명 길이 제한 (31자)
    if len(cleaned) > 31:
        cleaned = cleaned[:28] + "..."
    
    # 빈 문자열이면 기본값 사용
    if not cleaned.strip():
        cleaned = "Sheet"
    
    return cleaned

def merge_excel_files_smart(folder_path, output_file='통합파일.xlsx'):
    """
    폴더의 모든 엑셀 파일을 스마트하게 통합 (버전 관리 포함)
    """
    try:
        # 엑셀 파일들 찾기
        excel_files = []
        for ext in ['*.xlsx', '*.xls']:
            excel_files.extend(glob.glob(os.path.join(folder_path, ext)))
        
        # 출력 파일은 제외
        output_path = os.path.join(folder_path, output_file)
        excel_files = [f for f in excel_files if os.path.abspath(f) != os.path.abspath(output_path)]
        
        if not excel_files:
            print("❌ 폴더에 엑셀 파일이 없습니다.")
            return False
        
        print(f"📂 발견된 엑셀 파일: {len(excel_files)}개")
        
        # 최신 버전 파일들만 선택
        latest_files, version_info = find_latest_versions(excel_files)
        
        print("\n📋 버전 관리 결과:")
        print("-" * 60)
        for info in version_info:
            if info['old_versions']:
                print(f"📁 {info['base_name']} (버전 관리됨)")
                print(f"   ✅ 사용: {info['latest']}")
                for old in info['old_versions']:
                    print(f"   ❌ 제외: {old} (구버전)")
            else:
                print(f"📄 {info['latest']} (단일 버전)")
        
        print(f"\n🔄 총 {len(latest_files)}개 파일의 시트들을 통합합니다...")
        print("-" * 60)
        
        # 최소 하나의 시트가 있는지 확인할 변수
        valid_sheets_found = False
        
        # 엑셀 writer 객체 생성
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            processed_files = 0
            processed_sheets = 0
            
            for file_info in latest_files:
                try:
                    file_path = file_info['path']
                    filename = file_info['filename']
                    base_name, version = get_base_filename_and_version(filename)
                    
                    print(f"\n📄 처리 중: {filename}")
                    
                    # 엑셀 파일의 모든 시트 읽기 (안전한 방법)
                    try:
                        # 권한 확인
                        has_permission, permission_msg = check_file_permissions(file_path)
                        if not has_permission:
                            print(f"  ⚠️ 권한 문제: {permission_msg}")
                            
                            # 권한 수정 시도
                            fixed, fix_msg = fix_file_permissions(file_path)
                            if fixed:
                                print(f"  🔧 권한 수정: {fix_msg}")
                            else:
                                print(f"  ❌ 권한 수정 실패: {fix_msg}")
                                continue
                        
                        # 안전한 방법으로 시트 목록 가져오기
                        try:
                            import openpyxl
                            from io import BytesIO
                            
                            # 파일을 메모리로 읽기
                            with open(file_path, 'rb') as f:
                                file_bytes = f.read()
                            
                            # openpyxl로 워크북 열기 (스타일 무시)
                            wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
                            sheet_names = wb.sheetnames
                            
                        except Exception as e:
                            print(f"  ⚠️ 시트 목록 읽기 오류: {e}")
                            # 대안: pandas로 시트 목록 가져오기
                            try:
                                excel_data = pd.ExcelFile(file_path)
                                sheet_names = excel_data.sheet_names
                            except Exception as e2:
                                print(f"  ❌ 시트 목록 읽기 실패: {e2}")
                                continue
                        
                    except Exception as e:
                        print(f"  ❌ 파일 읽기 오류: {e}")
                        continue
                    
                    for sheet_name in sheet_names:
                        try:
                            # 안전한 시트 읽기
                            df = safe_read_excel(file_path, sheet_name)
                            if df is None:
                                print(f"  ❌ 시트 읽기 실패: {sheet_name}")
                                continue
                            
                            # 빈 시트 건너뛰기
                            if df.empty or len(df) == 0:
                                print(f"  ⚠️ 빈 시트: {sheet_name}")
                                continue
                            
                            # 모든 값이 None인 행 제거
                            df = df.dropna(how='all')
                            if df.empty or len(df) == 0:
                                print(f"  ⚠️ 유효한 데이터가 없는 시트: {sheet_name}")
                                continue
                            
                            # 시트명 생성 규칙
                            if len(sheet_names) == 1 and sheet_name.lower() in ['sheet1', 'sheet', '시트1']:
                                # 시트가 1개이고 기본 시트명인 경우: 파일명_버전
                                if version != "0":
                                    new_sheet_name = f"{base_name}_{version}"
                                else:
                                    new_sheet_name = base_name
                            else:
                                # 여러 시트가 있거나 의미있는 시트명인 경우: 파일명_시트명_버전
                                if version != "0":
                                    new_sheet_name = f"{base_name}_{sheet_name}_{version}"
                                else:
                                    new_sheet_name = f"{base_name}_{sheet_name}"
                            
                            # 시트명 정리
                            new_sheet_name = clean_sheet_name(new_sheet_name)
                            
                            # 시트명 중복 방지
                            original_name = new_sheet_name
                            counter = 1
                            while new_sheet_name in writer.sheets:
                                new_sheet_name = f"{original_name}_{counter}"
                                counter += 1
                            
                            # 엑셀 시트로 저장
                            df.to_excel(writer, sheet_name=new_sheet_name, index=False)
                            
                            print(f"  ✅ {sheet_name} → {new_sheet_name} ({len(df):,}행 {len(df.columns)}열)")
                            processed_sheets += 1
                            valid_sheets_found = True
                            
                        except Exception as e:
                            print(f"  ❌ 시트 처리 오류: {sheet_name} - {e}")
                    
                    processed_files += 1
                    
                except Exception as e:
                    print(f"  ❌ 파일 처리 오류: {filename} - {e}")
        
        print("\n" + "-" * 60)
        if processed_sheets > 0 and valid_sheets_found:
            print(f"✅ 통합 완료!")
            print(f"📊 처리 결과:")
            print(f"   - 처리된 파일: {processed_files}개")
            print(f"   - 통합된 시트: {processed_sheets}개")
            print(f"   - 저장 위치: {output_path}")
            return True
        else:
            print("❌ 처리된 시트가 없습니다.")
            print("💡 가능한 원인:")
            print("   - 모든 시트가 비어있음")
            print("   - 파일 읽기 권한 문제")
            print("   - Excel 파일 형식 문제")
            print("   - 파일이 다른 프로그램에서 사용 중")
            return False
            
    except Exception as e:
        print(f"❌ 프로그램 오류: {e}")
        return False

def main():
    """메인 실행 함수"""
    try:
        # 프로그램 헤더
        print("=" * 70)
        print("🚀 스마트 엑셀 파일 통합기 with 버전 관리")
        print("📧 제작자: charmleader@gmail.com")
        print("⏰ 실행 시간:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        print("=" * 70)
        
        # 현재 스크립트가 있는 폴더
        current_dir = os.path.dirname(os.path.abspath(__file__))
        print(f"📁 작업 폴더: {current_dir}")
        
        # 엑셀 파일들 찾기
        excel_files = []
        for ext in ['*.xlsx', '*.xls']:
            excel_files.extend(glob.glob(os.path.join(current_dir, ext)))
        
        # 출력 파일은 제외
        output_file = "통합파일.xlsx"
        output_path = os.path.join(current_dir, output_file)
        excel_files = [f for f in excel_files if os.path.abspath(f) != os.path.abspath(output_path)]
        
        if not excel_files:
            print("\n❌ 현재 폴더에 엑셀 파일이 없습니다.")
            print("\n📋 사용법:")
            print("   1. 이 스크립트(.py 파일)와 같은 폴더에 엑셀 파일들을 넣으세요")
            print("   2. 스크립트를 더블클릭하여 실행하세요")
            print("   3. 버전이 있는 파일들은 자동으로 최신 버전만 선택됩니다")
            print("\n💡 버전 관리 예시:")
            print("   - report.xlsx, report-1.xlsx, report-2.xlsx → report-2.xlsx만 사용")
            print("   - data_v1.xlsx, data_v2.xlsx → data_v2.xlsx만 사용")
            print("\n📊 시트명 생성 규칙:")
            print("   - 단일 시트: 파일명_버전")
            print("   - 여러 시트: 파일명_시트명_버전")
        else:
            print(f"\n🎯 작업 시작! 엑셀 파일 {len(excel_files)}개 발견")
            
            # 파일 통합 실행
            success = merge_excel_files_smart(current_dir, output_file)
            
            if success:
                print(f"\n🎉 작업 완료!")
                print(f"📊 결과 파일: {output_file}")
                print(f"💡 시트명 규칙: 파일명_시트명_버전 형태로 생성되었습니다!")
                print(f"🔄 구버전 파일들은 자동으로 제외되었습니다!")
            else:
                print(f"\n❌ 작업 실패! 위의 오류 메시지를 확인해주세요.")
        
        print("\n" + "=" * 70)
        print("💌 문의 및 피드백: charmleader@gmail.com")
        print("🔄 새로운 파일 추가 후 다시 실행하면 자동으로 반영됩니다!")
        
    except Exception as e:
        print(f"\n❌ 심각한 오류 발생: {e}")
        print("💌 이 오류를 charmleader@gmail.com으로 신고해주세요.")
    
    finally:
        input("\n✨ Enter 키를 눌러 프로그램을 종료하세요...")

if __name__ == "__main__":
    main()