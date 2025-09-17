import pandas as pd
import os
import glob
from pathlib import Path
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_education_name_and_date(df):
    """
    첫 번째 행에서 교육명과 날짜를 추출
    """
    if df.empty or len(df) < 1:
        return None, None, None
    
    first_row = df.iloc[0]
    education_name = None
    date_str = None
    time_str = None
    
    # 첫 번째 행의 모든 셀을 확인하여 교육명과 날짜 찾기
    for cell_value in first_row:
        if pd.isna(cell_value):
            continue
            
        cell_str = str(cell_value).strip()
        
        # 교육명 패턴 찾기 (한글이 포함된 경우)
        if not education_name and re.search(r'[가-힣]', cell_str) and len(cell_str) > 2:
            education_name = cell_str
        
        # 날짜 패턴 찾기 (MM-DD, MM/DD, MM.DD 형식)
        date_patterns = [
            r'(\d{1,2})[-/.](\d{1,2})',  # MM-DD, MM/DD, MM.DD
            r'(\d{1,2})월\s*(\d{1,2})일',  # MM월 DD일
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, cell_str)
            if match and not date_str:
                month, day = match.groups()
                date_str = f"{month.zfill(2)}-{day.zfill(2)}"
                break
        
        # 시간 패턴 찾기 (HH:MM 형식)
        time_match = re.search(r'(\d{1,2}):(\d{2})', cell_str)
        if time_match and not time_str:
            hour, minute = time_match.groups()
            time_str = f"{hour.zfill(2)}-{minute}"
    
    return education_name, date_str, time_str

def clean_sheet_name(name):
    """엑셀 시트명으로 사용할 수 없는 문자들 정리"""
    if not name:
        return "Sheet"
    
    # 엑셀에서 허용하지 않는 문자들 제거: [ ] : * ? / \
    invalid_chars = r'[\[\]:*?/\\]'
    cleaned = re.sub(invalid_chars, '_', str(name))
    
    # 시트명 길이 제한 (31자)
    if len(cleaned) > 31:
        cleaned = cleaned[:28] + "..."
    
    # 빈 문자열이면 기본값 사용
    if not cleaned.strip():
        cleaned = "Sheet"
    
    return cleaned

def process_sheet_data(df):
    """
    시트 데이터 처리: 첫 두 행 삭제, 취소 행 처리, 대기 학생 처리
    """
    if df.empty or len(df) < 3:
        return None
    
    # 첫 번째와 두 번째 행 삭제
    df = df.iloc[2:].reset_index(drop=True)
    
    if df.empty:
        return None
    
    # 컬럼명 정리 (첫 번째 행을 헤더로 사용)
    if len(df) > 0:
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
    
    # 빈 행 제거
    df = df.dropna(how='all')
    
    if df.empty:
        return None
    
    # 취소 행과 일반 행 분리
    cancelled_rows = []
    normal_rows = []
    
    for idx, row in df.iterrows():
        # 비고 열에서 '취소' 확인
        is_cancelled = False
        for col in df.columns:
            if pd.notna(row[col]) and '취소' in str(row[col]):
                is_cancelled = True
                break
        
        if is_cancelled:
            cancelled_rows.append(row)
        else:
            normal_rows.append(row)
    
    # 일반 행들로 DataFrame 재구성
    if normal_rows:
        df_normal = pd.DataFrame(normal_rows).reset_index(drop=True)
    else:
        df_normal = pd.DataFrame()
    
    # 취소 행들로 DataFrame 재구성
    if cancelled_rows:
        df_cancelled = pd.DataFrame(cancelled_rows).reset_index(drop=True)
    else:
        df_cancelled = pd.DataFrame()
    
    return df_normal, df_cancelled

def update_waitlist_status(df):
    """
    대기 학생 상태 업데이트
    """
    if df.empty:
        return df
    
    # 상태 열 찾기
    status_col = None
    for col in df.columns:
        if '상태' in str(col) or 'status' in str(col).lower():
            status_col = col
            break
    
    if status_col is None:
        return df
    
    # 대기 상태 업데이트
    for idx, row in df.iterrows():
        status_value = str(row[status_col]) if pd.notna(row[status_col]) else ""
        
        # 대기 번호 추출
        wait_match = re.search(r'대기(\d+)', status_value)
        if wait_match:
            wait_num = wait_match.group(1)
            df.at[idx, status_col] = f"Applied (대기{wait_num})"
    
    return df

def sort_dataframe(df):
    """
    데이터프레임 정렬
    1. 상태 – Descending
    2. 지역 – Ascending  
    3. 학교분류 – Descending
    4. 학교명 – Ascending
    5. 학년 – Ascending
    6. 반 – Ascending
    7. 번호 – Ascending
    """
    if df.empty:
        return df
    
    # 정렬할 컬럼들 찾기
    sort_columns = []
    sort_orders = []
    
    column_mapping = {
        '상태': ('상태', False),  # Descending
        '지역': ('지역', True),   # Ascending
        '학교분류': ('학교분류', False),  # Descending
        '학교명': ('학교명', True),  # Ascending
        '학년': ('학년', True),   # Ascending
        '반': ('반', True),       # Ascending
        '번호': ('번호', True)    # Ascending
    }
    
    for col_name, (ascending, is_asc) in column_mapping.items():
        for df_col in df.columns:
            if col_name in str(df_col):
                sort_columns.append(df_col)
                sort_orders.append(is_asc)
                break
    
    if sort_columns:
        df = df.sort_values(by=sort_columns, ascending=sort_orders, na_position='last')
    
    return df

def add_grade_class_number_column(df):
    """
    학년-반-번호 조합 열 추가
    """
    if df.empty:
        return df
    
    # 학년, 반, 번호 열 찾기
    grade_col = None
    class_col = None
    number_col = None
    
    for col in df.columns:
        col_str = str(col)
        if '학년' in col_str:
            grade_col = col
        elif '반' in col_str and '학년' not in col_str:
            class_col = col
        elif '번호' in col_str:
            number_col = col
    
    if grade_col and class_col and number_col:
        # 조합 열 생성
        def combine_grade_class_number(row):
            grade = str(row[grade_col]) if pd.notna(row[grade_col]) else ""
            class_num = str(row[class_col]) if pd.notna(row[class_col]) else ""
            number = str(row[number_col]) if pd.notna(row[number_col]) else ""
            
            if grade and class_num and number:
                return f"{grade}-{class_num}-{number}"
            return ""
        
        df['학년-반-번호'] = df.apply(combine_grade_class_number, axis=1)
    
    return df

def add_class_info_columns(df, class_info_df):
    """
    수업 정보 열 추가
    """
    if df.empty or class_info_df.empty:
        return df
    
    # 기본 열들 추가
    new_columns = ['수업일', '시작', '종료', '주강사', '보조강사', '장소', '모니터']
    for col in new_columns:
        df[col] = ""
    
    # 수업 정보가 있으면 매핑 (간단한 예시)
    # 실제로는 교육명이나 다른 키를 기준으로 매핑해야 함
    if '교육명' in class_info_df.columns:
        for idx, row in df.iterrows():
            # 여기서는 간단히 첫 번째 수업 정보를 사용
            if not class_info_df.empty:
                first_class = class_info_df.iloc[0]
                for col in new_columns:
                    if col in class_info_df.columns:
                        df.at[idx, col] = str(first_class[col]) if pd.notna(first_class[col]) else ""
    
    return df

def apply_strikethrough_to_cancelled_rows(worksheet, cancelled_df, start_row):
    """
    취소된 행에 취소선 적용
    """
    if cancelled_df.empty:
        return
    
    for idx, row in cancelled_df.iterrows():
        row_num = start_row + idx + 1
        for col_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_idx)
            cell.font = Font(strike=True)
            cell.value = value

def merge_excel_files_advanced(folder_path, class_info_file=None, output_file='통합파일_고급.xlsx'):
    """
    고급 엑셀 파일 통합
    """
    try:
        # 엑셀 파일들 찾기
        excel_files = []
        for ext in ['*.xlsx', '*.xls']:
            excel_files.extend(glob.glob(os.path.join(folder_path, ext)))
        
        # 출력 파일과 수업 정보 파일 제외
        output_path = os.path.join(folder_path, output_file)
        excel_files = [f for f in excel_files if os.path.abspath(f) != os.path.abspath(output_path)]
        if class_info_file:
            excel_files = [f for f in excel_files if os.path.abspath(f) != os.path.abspath(class_info_file)]
        
        if not excel_files:
            print("❌ 폴더에 엑셀 파일이 없습니다.")
            return False
        
        print(f"📂 발견된 엑셀 파일: {len(excel_files)}개")
        
        # 수업 정보 파일 읽기
        class_info_df = pd.DataFrame()
        if class_info_file and os.path.exists(class_info_file):
            try:
                class_info_df = pd.read_excel(class_info_file)
                print(f"📋 수업 정보 파일 로드: {class_info_file}")
            except Exception as e:
                print(f"⚠️ 수업 정보 파일 읽기 실패: {e}")
        
        # 교육명별로 그룹화하여 처리
        education_groups = {}
        
        for file_path in excel_files:
            try:
                filename = os.path.basename(file_path)
                print(f"\n📄 처리 중: {filename}")
                
                # 엑셀 파일의 모든 시트 읽기
                excel_data = pd.ExcelFile(file_path)
                
                for sheet_name in excel_data.sheet_names:
                    try:
                        # 시트 읽기
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        
                        # 빈 시트 건너뛰기
                        if df.empty:
                            print(f"  ⚠️ 빈 시트: {sheet_name}")
                            continue
                        
                        # 교육명과 날짜 추출
                        education_name, date_str, time_str = extract_education_name_and_date(df)
                        
                        if not education_name:
                            print(f"  ⚠️ 교육명을 찾을 수 없음: {sheet_name}")
                            continue
                        
                        # 시트 데이터 처리
                        result = process_sheet_data(df)
                        if result is None:
                            print(f"  ⚠️ 처리할 데이터가 없음: {sheet_name}")
                            continue
                        
                        df_normal, df_cancelled = result
                        
                        if df_normal.empty:
                            print(f"  ⚠️ 유효한 데이터가 없음: {sheet_name}")
                            continue
                        
                        # 대기 학생 상태 업데이트
                        df_normal = update_waitlist_status(df_normal)
                        
                        # 정렬
                        df_normal = sort_dataframe(df_normal)
                        
                        # 학년-반-번호 열 추가
                        df_normal = add_grade_class_number_column(df_normal)
                        
                        # 수업 정보 열 추가
                        df_normal = add_class_info_columns(df_normal, class_info_df)
                        
                        # 시트명 생성
                        if date_str and time_str:
                            sheet_name_new = f"{education_name}_{date_str}_{time_str}"
                        elif date_str:
                            sheet_name_new = f"{education_name}_{date_str}"
                        else:
                            sheet_name_new = education_name
                        
                        sheet_name_new = clean_sheet_name(sheet_name_new)
                        
                        # 교육명별로 그룹화
                        if education_name not in education_groups:
                            education_groups[education_name] = []
                        
                        education_groups[education_name].append({
                            'sheet_name': sheet_name_new,
                            'data': df_normal,
                            'cancelled_data': df_cancelled,
                            'date_str': date_str,
                            'time_str': time_str,
                            'original_file': filename
                        })
                        
                        print(f"  ✅ {sheet_name} → {sheet_name_new} ({len(df_normal):,}행)")
                        
                    except Exception as e:
                        print(f"  ❌ 시트 처리 오류: {sheet_name} - {e}")
                
            except Exception as e:
                print(f"  ❌ 파일 처리 오류: {filename} - {e}")
        
        # 각 교육명별로 최신 버전만 선택하고 통합
        final_sheets = {}
        
        for education_name, sheets in education_groups.items():
            if not sheets:
                continue
            
            # 날짜별로 그룹화
            date_groups = {}
            for sheet_info in sheets:
                date_key = sheet_info['date_str'] or 'no_date'
                if date_key not in date_groups:
                    date_groups[date_key] = []
                date_groups[date_key].append(sheet_info)
            
            # 각 날짜별로 최신 버전 선택
            for date_key, date_sheets in date_groups.items():
                if len(date_sheets) == 1:
                    # 단일 시트
                    final_sheets[date_sheets[0]['sheet_name']] = date_sheets[0]
                else:
                    # 여러 시트가 있는 경우 버전 추가
                    for i, sheet_info in enumerate(date_sheets, 1):
                        if i > 1:
                            sheet_info['sheet_name'] = f"{sheet_info['sheet_name']}*{i-1}"
                        final_sheets[sheet_info['sheet_name']] = sheet_info
        
        # 엑셀 파일 생성
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, sheet_info in final_sheets.items():
                # 일반 데이터 저장
                sheet_info['data'].to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 취소된 데이터가 있으면 추가
                if not sheet_info['cancelled_data'].empty:
                    # 취소된 데이터를 맨 아래에 추가
                    combined_data = pd.concat([
                        sheet_info['data'],
                        sheet_info['cancelled_data']
                    ], ignore_index=True)
                    
                    # 다시 저장
                    combined_data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 취소선 적용을 위해 openpyxl로 다시 열어서 스타일 적용
        wb = openpyxl.load_workbook(output_path)
        
        for sheet_name, sheet_info in final_sheets.items():
            if not sheet_info['cancelled_data'].empty:
                ws = wb[sheet_name]
                normal_rows = len(sheet_info['data'])
                apply_strikethrough_to_cancelled_rows(ws, sheet_info['cancelled_data'], normal_rows)
        
        wb.save(output_path)
        
        print(f"\n✅ 통합 완료!")
        print(f"📊 처리 결과:")
        print(f"   - 통합된 시트: {len(final_sheets)}개")
        print(f"   - 저장 위치: {output_path}")
        
        return True
        
    except Exception as e:
        print(f"❌ 프로그램 오류: {e}")
        return False

def main():
    """메인 실행 함수"""
    try:
        # 프로그램 헤더
        print("=" * 70)
        print("🚀 고급 엑셀 파일 통합기")
        print("📧 제작자: charmleader@gmail.com")
        print("⏰ 실행 시간:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        print("=" * 70)
        
        # 현재 스크립트가 있는 폴더
        current_dir = os.path.dirname(os.path.abspath(__file__))
        print(f"📁 작업 폴더: {current_dir}")
        
        # 수업 정보 파일 찾기
        class_info_files = glob.glob(os.path.join(current_dir, "*수업정보*.xlsx"))
        class_info_file = class_info_files[0] if class_info_files else None
        
        if class_info_file:
            print(f"📋 수업 정보 파일 발견: {os.path.basename(class_info_file)}")
        else:
            print("⚠️ 수업 정보 파일을 찾을 수 없습니다. (파일명에 '수업정보'가 포함된 xlsx 파일)")
        
        # 파일 통합 실행
        success = merge_excel_files_advanced(current_dir, class_info_file)
        
        if success:
            print(f"\n🎉 작업 완료!")
            print(f"📊 결과 파일: 통합파일_고급.xlsx")
            print(f"💡 처리된 기능:")
            print(f"   - 빈 시트 제외")
            print(f"   - 교육명 기반 시트명 생성")
            print(f"   - 날짜/시간 정보 포함")
            print(f"   - 취소 행 취소선 처리")
            print(f"   - 대기 학생 상태 업데이트")
            print(f"   - 데이터 정렬")
            print(f"   - 학년-반-번호 조합 열 추가")
            print(f"   - 수업 정보 열 추가")
        else:
            print(f"\n❌ 작업 실패! 위의 오류 메시지를 확인해주세요.")
        
        print("\n" + "=" * 70)
        print("💌 문의 및 피드백: charmleader@gmail.com")
        
    except Exception as e:
        print(f"\n❌ 심각한 오류 발생: {e}")
        print("💌 이 오류를 charmleader@gmail.com으로 신고해주세요.")
    
    finally:
        input("\n✨ Enter 키를 눌러 프로그램을 종료하세요...")

if __name__ == "__main__":
    main()