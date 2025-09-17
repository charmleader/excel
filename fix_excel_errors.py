#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 파일 오류 수정 도구
openpyxl 스타일 오류 및 기타 문제를 해결합니다.
"""

import os
import shutil
import tempfile
from pathlib import Path
import pandas as pd
import openpyxl
from io import BytesIO

def fix_excel_file(input_path, output_path=None):
    """Excel 파일의 스타일 오류를 수정합니다."""
    try:
        if output_path is None:
            output_path = input_path.replace('.xlsx', '_fixed.xlsx')
        
        print(f"🔧 파일 수정 중: {os.path.basename(input_path)}")
        
        # 1단계: 원본 파일을 메모리로 읽기
        with open(input_path, 'rb') as f:
            file_bytes = f.read()
        
        # 2단계: openpyxl로 워크북 열기 (스타일 무시)
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        
        # 3단계: 새 워크북 생성
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # 기본 시트 제거
        
        # 4단계: 각 시트를 새 워크북에 복사
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                new_ws = new_wb.create_sheet(title=sheet_name)
                
                # 데이터만 복사 (스타일 제외)
                for row in ws.iter_rows(values_only=True):
                    new_ws.append(row)
                
                print(f"  ✅ 시트 복사 완료: {sheet_name}")
                
            except Exception as e:
                print(f"  ⚠️ 시트 복사 실패: {sheet_name} - {e}")
                continue
        
        # 5단계: 수정된 파일 저장
        new_wb.save(output_path)
        print(f"✅ 수정 완료: {os.path.basename(output_path)}")
        return True
        
    except Exception as e:
        print(f"❌ 파일 수정 실패: {e}")
        return False

def batch_fix_excel_files(folder_path):
    """폴더 내 모든 Excel 파일을 일괄 수정합니다."""
    excel_files = []
    for ext in ['*.xlsx', '*.xls']:
        excel_files.extend(Path(folder_path).glob(ext))
    
    if not excel_files:
        print("❌ 폴더에 Excel 파일이 없습니다.")
        return
    
    print(f"📁 발견된 Excel 파일: {len(excel_files)}개")
    print("-" * 50)
    
    success_count = 0
    for file_path in excel_files:
        if fix_excel_file(str(file_path)):
            success_count += 1
        print()
    
    print("=" * 50)
    print(f"✅ 수정 완료: {success_count}/{len(excel_files)}개 파일")

def main():
    """메인 실행 함수"""
    print("=" * 60)
    print("🔧 Excel 파일 오류 수정 도구")
    print("📧 제작자: charmleader@gmail.com")
    print("=" * 60)
    
    current_dir = os.getcwd()
    print(f"📁 작업 폴더: {current_dir}")
    
    # 현재 폴더의 Excel 파일들 수정
    batch_fix_excel_files(current_dir)
    
    print("\n💡 사용법:")
    print("1. 이 스크립트를 Excel 파일이 있는 폴더에서 실행하세요")
    print("2. 수정된 파일은 '_fixed.xlsx' 접미사가 붙습니다")
    print("3. 원본 파일은 그대로 유지됩니다")
    
    input("\n✨ Enter 키를 눌러 종료하세요...")

if __name__ == "__main__":
    main()
